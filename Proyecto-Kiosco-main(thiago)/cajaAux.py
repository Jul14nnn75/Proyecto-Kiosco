import tkinter as tk
from tkinter import messagebox
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

# -----------------------------------------
# 🔸 SISTEMA DE EMAIL INTEGRADO
# -----------------------------------------
class EnviadorEmail:
    def __init__(self):
        # Configuración de email (misma que en reporte.py)
        self.servidor_smtp = "smtp.gmail.com"
        self.puerto_smtp = 587
        self.email_remitente = "Software.Kiosco@gmail.com"
        self.password_email = "feyu zacq rcsc jedy"  # Contraseña de aplicación de Gmail

    def enviar_email(self, email_destino, asunto, mensaje, archivo_adjunto=None):
        """
        Envía un email con opción de adjuntar archivo
        """
        try:
            # Crear mensaje
            mensaje_email = MIMEMultipart()
            mensaje_email['From'] = self.email_remitente
            mensaje_email['To'] = email_destino
            mensaje_email['Subject'] = asunto

            # Cuerpo del mensaje
            mensaje_email.attach(MIMEText(mensaje, 'plain'))

            # Adjuntar archivo si se proporciona
            if archivo_adjunto and os.path.exists(archivo_adjunto):
                with open(archivo_adjunto, "rb") as archivo:
                    parte = MIMEBase('application', 'octet-stream')
                    parte.set_payload(archivo.read())
                    
                encoders.encode_base64(parte)
                parte.add_header(
                    'Content-Disposition',
                    f'attachment; filename="{os.path.basename(archivo_adjunto)}"'
                )
                mensaje_email.attach(parte)

            # Enviar email
            with smtplib.SMTP(self.servidor_smtp, self.puerto_smtp) as servidor:
                servidor.starttls()
                servidor.login(self.email_remitente, self.password_email)
                servidor.send_message(mensaje_email)

            return True, "Email enviado correctamente"
            
        except Exception as e:
            return False, f"Error enviando email: {str(e)}"

# Instancia global para usar en toda la aplicación
enviador_email = EnviadorEmail()

# -----------------------------------------
# 🔸 VENTANA EMERGENTE REUTILIZABLE: Apertura / Retiro / Ingreso / Gasto
# -----------------------------------------
class VentanaEmergenteCaja(tk.Toplevel):
    def __init__(self, master, titulo, tipo, color, callback=None):
        super().__init__(master)
        self.title(titulo)
        self.geometry("400x300")
        self.transient(master)
        self.grab_set()
        self.configure(bg=master.bg_color)
        self.callback = callback
        self.tipo = tipo
        self.master = master
        self.color = color

        self.centrar_ventana(400, 300)

        marco = tk.Frame(self, bg=master.card_color, padx=20, pady=20)
        marco.pack(fill="both", expand=True, padx=10, pady=10)

        tk.Label(marco, text=f"💰 {titulo.upper()}",
                 font=("Arial", 16, "bold"),
                 bg=master.card_color, fg=master.text_color).pack(pady=(0, 15))

        if tipo not in ("apertura", "gasto"):
            saldo = f"{master.saldo_sistema:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            marco_saldo = tk.Frame(marco, bg="#f8f9fa", relief="solid", bd=1, padx=10, pady=6)
            marco_saldo.pack(fill="x", pady=(0, 15))
            tk.Label(marco_saldo, text=f"Saldo actual: ${saldo}", font=("Arial", 11, "bold"),
                     bg="#f8f9fa", fg=master.text_color).pack()

        campos = tk.Frame(marco, bg=master.card_color)
        campos.pack(fill="x", pady=10)

        if tipo != "apertura":
            tk.Label(campos, text=f"Motivo del {tipo}:", font=("Arial", 11),
                     bg=master.card_color, fg=master.text_color).pack(anchor="w")
            self.entrada_motivo = tk.Entry(campos, font=("Arial", 12), relief="solid", bd=1)
            self.entrada_motivo.pack(fill="x", pady=(5, 15), ipady=5)
        else:
            self.entrada_motivo = None

        tk.Label(campos, text=f"Monto a {tipo}:", font=("Arial", 11),
                 bg=master.card_color, fg=master.text_color).pack(anchor="w")
        self.entrada_monto = tk.Entry(campos, font=("Arial", 14, "bold"), relief="solid", bd=1, justify="center")
        self.entrada_monto.pack(fill="x", pady=(5, 0), ipady=5)
        self.entrada_monto.focus()

        botones = tk.Frame(marco, bg=master.card_color)
        botones.pack(fill="x", pady=(20, 0))

        tk.Button(botones, text="Aceptar", command=self.aceptar,
                  bg=color, fg="white", font=("Arial", 11, "bold"),
                  relief="flat", padx=20, pady=10).pack(side="right", padx=5)
        tk.Button(botones, text="Cancelar", command=self.destroy,
                  bg=master.danger_color, fg="white", font=("Arial", 11),
                  relief="flat", padx=20, pady=10).pack(side="right", padx=5)

        self.bind("<Return>", lambda e: self.aceptar())
        self.bind("<Escape>", lambda e: self.destroy())

    def centrar_ventana(self, ancho, alto):
        x = (self.winfo_screenwidth() - ancho) // 2
        y = (self.winfo_screenheight() - alto) // 2
        self.geometry(f"{ancho}x{alto}+{x}+{y}")

    def aceptar(self):
        texto_monto = self.entrada_monto.get().strip().replace(".", "").replace(",", ".")
        try:
            monto = float(texto_monto)
        except:
            messagebox.showerror("Error", "Ingrese un monto válido.")
            return

        motivo = self.entrada_motivo.get().strip() if self.entrada_motivo else "(Apertura)"

        if self.tipo == "ingreso":
            self.master.saldo_sistema += monto
            self.master.ingreso_actual = (motivo, monto)
        elif self.tipo == "retiro":
            if monto > self.master.saldo_sistema:
                messagebox.showerror("Error", "El monto excede el saldo disponible.")
                return
            self.master.saldo_sistema -= monto
            self.master.retiro_actual = (motivo, monto)
        elif self.tipo == "gasto":
            if not hasattr(self.master, "gastos"):
                self.master.gastos = []
            self.master.gastos.append((motivo, monto))
        elif self.tipo == "apertura":
            self.master.monto_apertura = monto
            self.master.saldo_sistema = monto

        if self.callback:
            self.callback(monto, motivo)

        self.master._actualizar_estado()
        self.master._guardar_estado()
        self.destroy()


# -----------------------------------------
# 🔒 VENTANA DE CIERRE DE CAJA CON EMAIL AUTOMÁTICO
# -----------------------------------------
def ventana_cierre_caja(master):
    ventana = tk.Toplevel(master)
    ventana.title("🔒 Cierre de Caja")
    ventana.geometry("500x500")
    ventana.configure(bg=master.bg_color)
    ventana.transient(master)
    ventana.grab_set()
    ventana.lift()
    ventana.attributes("-topmost", True)

    marco = tk.Frame(ventana, bg=master.card_color, padx=25, pady=25)
    marco.pack(fill="both", expand=True, padx=15, pady=15)

    tk.Label(marco, text="🧾 RESUMEN DE CAJA", font=("Arial", 15, "bold"),
             bg=master.card_color, fg=master.text_color).pack(pady=(0, 15))

    saldo_sistema = getattr(master, "saldo_sistema", 0.0)

    tk.Label(marco, text="💻 Saldo del Sistema:", font=("Arial", 11, "bold"),
             bg=master.card_color, fg=master.text_color).pack(anchor="w")
    tk.Label(marco, text=f"${saldo_sistema:.2f}", font=("Arial", 11),
             bg=master.card_color, fg=master.text_color).pack(anchor="w", pady=(0, 10))

    tk.Label(marco, text="💵 Importe de Caja:", font=("Arial", 11, "bold"),
             bg=master.card_color, fg=master.text_color).pack(anchor="w")
    entrada_caja = tk.Entry(marco, font=("Arial", 12), relief="solid", bd=1)
    entrada_caja.pack(fill="x", pady=(5, 10))

    tk.Label(marco, text="🏦 Importe Transferencia:", font=("Arial", 11, "bold"),
             bg=master.card_color, fg=master.text_color).pack(anchor="w")
    entrada_transferencia = tk.Entry(marco, font=("Arial", 12), relief="solid", bd=1)
    entrada_transferencia.pack(fill="x", pady=(5, 10))

    etiqueta_diferencia = tk.Label(marco, text="Diferencia: $0.00", font=("Arial", 12, "bold"),
                       bg=master.card_color, fg=master.text_color)
    etiqueta_diferencia.pack(pady=(10, 5))

    # 🔥 EMAIL DESTINO FIJO
    marco_email = tk.LabelFrame(marco, text="📧 Envío Automático de Reporte", 
                               bg=master.card_color, fg=master.text_color,
                               font=("Arial", 11, "bold"), pady=10)
    marco_email.pack(fill="x", pady=(20, 10))

    tk.Label(marco_email, text="El reporte Excel se enviará automáticamente a:", 
             font=("Arial", 10), bg=master.card_color, fg=master.text_color).pack(anchor="w")
    
    email_destino = "thiagovillalba63@gmail.com"  # ⚠️ CAMBIA ESTE EMAIL AL QUE QUIERAS
    tk.Label(marco_email, text=f"📨 {email_destino}", 
             font=("Arial", 10, "bold"), bg=master.card_color, fg="#2c3e50").pack(anchor="w", pady=(5, 0))

    def calcular_diferencia(*args):
        try:
            caja = float(entrada_caja.get().replace(",", ".") or 0)
            transferencia = float(entrada_transferencia.get().replace(",", ".") or 0)
            diferencia = caja + transferencia - saldo_sistema
            etiqueta_diferencia.config(text=f"Diferencia: ${diferencia:.2f}",
                           fg="green" if diferencia >= 0 else "red")
        except ValueError:
            etiqueta_diferencia.config(text="Diferencia: $0.00", fg="black")

    entrada_caja.bind("<KeyRelease>", calcular_diferencia)
    entrada_transferencia.bind("<KeyRelease>", calcular_diferencia)

    botones = tk.Frame(marco, bg=master.card_color)
    botones.pack(fill="x", pady=(20, 0))

    tk.Button(botones, text="Cerrar Caja y Enviar Email", font=("Arial", 11, "bold"), bg="#27ae60", fg="white",
              command=lambda: confirmar_cierre(master, ventana, entrada_caja, entrada_transferencia, email_destino)).pack(side="right", padx=5)
    
    tk.Button(botones, text="Cancelar", font=("Arial", 11), bg="#e74c3c", fg="white",
              command=ventana.destroy).pack(side="right", padx=5)


def confirmar_cierre(master, ventana, entrada_caja, entrada_transferencia, email_destino):
    try:
        caja = float(entrada_caja.get().replace(",", ".") or 0)
        transferencia = float(entrada_transferencia.get().replace(",", ".") or 0)
    except ValueError:
        messagebox.showerror("Error", "Ingrese valores numéricos válidos.")
        return

    saldo_sistema = getattr(master, "saldo_sistema", 0)
    diferencia = caja + transferencia - saldo_sistema

    ingreso = getattr(master, "ingreso_actual", ("", 0))
    retiro = getattr(master, "retiro_actual", ("", 0))
    gastos = getattr(master, "gastos", [])

    # Guardar cierre localmente
    archivo_excel = guardar_cierre(saldo_sistema, caja, transferencia, diferencia, ingreso, retiro, gastos)

    # 🔥 ENVIAR EMAIL AUTOMÁTICAMENTE
    try:
        enviar_email_con_archivo(email_destino, archivo_excel)
        messagebox.showinfo("Éxito", f"✅ Cierre guardado y enviado a:\n{email_destino}")
    except Exception as e:
        messagebox.showerror("Error Email", f"✅ Cierre guardado\n❌ Error enviando email:\n{str(e)}")

    # Resetear todo
    master.monto_apertura = None
    master.saldo_sistema = 0
    master.ingreso_actual = ("", 0)
    master.retiro_actual = ("", 0)
    master.gastos = []
    master._actualizar_estado()
    master._guardar_estado()

    ventana.destroy()


# -----------------------------------------
# 📧 FUNCIÓN PARA ENVIAR EMAIL CON EXCEL
# -----------------------------------------
def enviar_email_con_archivo(email_destino, archivo_excel):
    """
    Envía el archivo Excel por email automáticamente
    """
    try:
        asunto = f"Cierre de Caja - {datetime.now().strftime('%d/%m/%Y')}"
        
        mensaje = f"""
        Buenos días,
        
        Adjunto el reporte de cierre de caja correspondiente a la fecha.
        
        Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}
        Archivo: {os.path.basename(archivo_excel)}
        
        Este es un mensaje automático del Sistema de Kiosco.
        
        Saludos cordiales,
        Sistema de Kiosco
        """
        
        # Usar el sistema de email
        exito, mensaje_resultado = enviador_email.enviar_email(
            email_destino=email_destino,
            asunto=asunto,
            mensaje=mensaje,
            archivo_adjunto=archivo_excel
        )
        
        if not exito:
            raise Exception(mensaje_resultado)
            
    except Exception as e:
        raise Exception(f"No se pudo enviar el email: {str(e)}")


# -----------------------------------------
# 📊 GUARDAR CIERRE DE CAJA (Excel)
# -----------------------------------------
def guardar_cierre(saldo_sistema, importe_caja, importe_transferencia, diferencia, ingreso, retiro, gastos):
    os.makedirs("data/cierres", exist_ok=True)
    ruta = os.path.join("data/cierres", "cierres.xlsx")

    if os.path.exists(ruta):
        wb = load_workbook(ruta)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen de Caja"

        ws.merge_cells("A1:L1")
        celda_titulo = ws["A1"]
        celda_titulo.value = "Resumen de Caja"
        celda_titulo.font = Font(size=16, bold=True, color="FFFFFF")
        celda_titulo.alignment = Alignment(horizontal="center", vertical="center")
        celda_titulo.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")

        encabezados = [
            "Fecha", "Saldo Sistema", "Importe Caja", "Importe Transferencia", "Diferencia",
            "Motivo Ingreso", "Ingreso", "Motivo Retiro", "Retiro", "Motivo Gasto", "Gasto", "Gasto 2..."
        ]
        ws.append([""] * len(encabezados))
        ws.append(encabezados)
        for col in range(1, len(encabezados) + 1):
            ws.column_dimensions[chr(64 + col)].width = 20

    fila = [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        f"{saldo_sistema:.2f}",
        f"{importe_caja:.2f}",
        f"{importe_transferencia:.2f}",
        f"{diferencia:.2f}",
        ingreso[0], f"{ingreso[1]:.2f}",
        retiro[0], f"{retiro[1]:.2f}",
    ]

    # Agregar todos los gastos por separado
    for g in gastos:
        fila.extend([g[0], f"{g[1]:.2f}"])

    ws.append(fila)
    wb.save(ruta)
    print(f"✅ Cierre guardado en {ruta}")
    return ruta